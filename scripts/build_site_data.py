#!/usr/bin/env python3
"""Build the static website data from the Excel workbook and statute document."""

from __future__ import annotations

import json
import re
import sys
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Any

try:
    import openpyxl
    from docx import Document
    from docx.oxml.table import CT_Tbl
    from docx.oxml.text.paragraph import CT_P
    from docx.table import Table
    from docx.text.paragraph import Paragraph
    from PIL import Image
except ImportError as exc:  # pragma: no cover - helpful for local usage
    print(f"Modulo mancante: {exc.name}", file=sys.stderr)
    print("Installa le dipendenze con: python3 -m pip install -r requirements.txt", file=sys.stderr)
    raise SystemExit(1)


ROOT = Path(__file__).resolve().parents[1]
EXCEL_PATH = ROOT / "calottino_categoria_gestione.xlsx"
STATUTE_PATH = ROOT / "statuto_calottino_categoria_avionici.docx"
LOGO_SOURCE_PATH = ROOT / "WhatsApp Image 2026-06-16 at 15.08.20.jpeg"
DOCS_DIR = ROOT / "docs"
DATA_DIR = DOCS_DIR / "data"
ASSETS_DIR = DOCS_DIR / "assets"


def clean_value(value: Any) -> Any:
    if isinstance(value, datetime):
        return value.strftime("%d/%m/%Y")
    if isinstance(value, date):
        return value.strftime("%d/%m/%Y")
    if isinstance(value, float) and value.is_integer():
        return int(value)
    if isinstance(value, str):
        normalized = value.replace("\xa0", " ").strip()
        return normalized if normalized else None
    return value


def present(value: Any) -> bool:
    return value is not None and value != ""


def number(value: Any, fallback: float = 0) -> float:
    value = clean_value(value)
    if value is None:
        return fallback
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        try:
            return float(value.replace(".", "").replace(",", "."))
        except ValueError:
            return fallback
    return fallback


def table_from_sheet(
    workbook: Any,
    sheet_name: str,
    header_row: int,
    required_headers: list[str] | None = None,
) -> dict[str, Any]:
    sheet = workbook[sheet_name]
    headers = [clean_value(sheet.cell(header_row, col).value) for col in range(1, sheet.max_column + 1)]
    headers = [str(header) for header in headers if present(header)]
    rows: list[dict[str, Any]] = []

    for row_idx in range(header_row + 1, sheet.max_row + 1):
        row: dict[str, Any] = {}
        for col_idx, header in enumerate(headers, start=1):
            row[header] = clean_value(sheet.cell(row_idx, col_idx).value)

        keys = required_headers or headers
        if any(present(row.get(key)) for key in keys):
            rows.append(row)

    return {
        "title": clean_value(sheet.cell(1, 1).value) or sheet_name,
        "sheet": sheet_name,
        "headers": headers,
        "rows": rows,
    }


def table_range(
    sheet: Any,
    title_cell: str,
    header_row: int,
    first_col: int,
    last_col: int,
    data_start: int,
    data_end: int,
) -> dict[str, Any]:
    headers = [
        str(clean_value(sheet.cell(header_row, col).value))
        for col in range(first_col, last_col + 1)
        if present(clean_value(sheet.cell(header_row, col).value))
    ]
    rows: list[dict[str, Any]] = []

    for row_idx in range(data_start, data_end + 1):
        row: dict[str, Any] = {}
        for offset, header in enumerate(headers):
            row[header] = clean_value(sheet.cell(row_idx, first_col + offset).value)
        if any(present(value) for value in row.values()):
            rows.append(row)

    return {
        "title": clean_value(sheet[title_cell].value),
        "headers": headers,
        "rows": rows,
    }


def lookup(rows: list[dict[str, Any]], label: str, value_key: str = "Importo") -> Any:
    for row in rows:
        if row.get("Voce") == label:
            return row.get(value_key)
    return None


def read_workbook() -> dict[str, Any]:
    workbook = openpyxl.load_workbook(EXCEL_PATH, data_only=True)

    people = table_from_sheet(
        workbook,
        "Persone_Quote",
        header_row=3,
        required_headers=["ID", "Nome / Cognome"],
    )
    purchases = table_from_sheet(
        workbook,
        "Acquisti_Patch",
        header_row=3,
        required_headers=["Data", "Fornitore", "Lotto / Descrizione", "Quantità", "Costo totale", "Pagato da"],
    )
    sales = table_from_sheet(
        workbook,
        "Vendite_Patch",
        header_row=3,
        required_headers=["Data", "Acquirente", "Quantità", "Ricavo", "Metodo pagamento", "Stato consegna"],
    )
    expenses = table_from_sheet(
        workbook,
        "Spese_Categoria",
        header_row=3,
        required_headers=["Data", "Categoria", "Descrizione", "Importo", "Pagato da"],
    )
    parameters = table_from_sheet(workbook, "Parametri", header_row=3, required_headers=["Voce"])

    dashboard_sheet = workbook["Dashboard"]
    flows = table_range(dashboard_sheet, "A13", 14, 1, 4, 15, 19)
    inventory = table_range(dashboard_sheet, "G13", 14, 7, 10, 15, 19)

    member_rows = people["rows"]
    paid_members = sum(1 for row in member_rows if row.get("Stato") == "Pagata")
    partial_members = sum(1 for row in member_rows if row.get("Stato") == "Parziale")
    due_total = sum(number(row.get("Quota dovuta")) for row in member_rows)
    paid_total = sum(number(row.get("Quota pagata")) for row in member_rows)
    remaining_total = max(due_total - paid_total, 0)

    flow_rows = flows["rows"]
    inventory_rows = inventory["rows"]
    summary = {
        "cashAvailable": number(lookup(flow_rows, "Cassa disponibile stimata")),
        "duesCollected": number(lookup(flow_rows, "Quote incassate")),
        "patchRevenue": number(lookup(flow_rows, "Ricavi patch incassati")),
        "patchPurchaseCost": number(lookup(flow_rows, "Costo acquisti patch")),
        "categoryExpenses": number(lookup(flow_rows, "Spese categoria")),
        "patchGrossProfit": number(lookup(flow_rows, "Utile lordo patch")),
        "patchPurchased": number(lookup(inventory_rows, "Patch acquistate", "Valore")),
        "patchSold": number(lookup(inventory_rows, "Patch vendute", "Valore")),
        "patchAvailable": number(lookup(inventory_rows, "Patch disponibili", "Valore")),
        "breakEvenPatches": number(lookup(inventory_rows, "Break-even acquisti", "Valore")),
        "patchUnitMargin": number(lookup(inventory_rows, "Margine unitario", "Valore")),
        "membersTotal": len(member_rows),
        "membersPaid": paid_members,
        "membersPartial": partial_members,
        "membersOpen": max(len(member_rows) - paid_members - partial_members, 0),
        "duesExpected": due_total,
        "duesRemaining": remaining_total,
        "duesCompletionRate": round((paid_total / due_total) * 100, 1) if due_total else 0,
        "purchasesCount": len(purchases["rows"]),
        "salesCount": len(sales["rows"]),
        "expensesCount": len(expenses["rows"]),
    }

    return {
        "workbookTitle": clean_value(dashboard_sheet["A1"].value),
        "workbookSubtitle": clean_value(dashboard_sheet["A2"].value),
        "summary": summary,
        "parameters": parameters,
        "people": people,
        "purchases": purchases,
        "sales": sales,
        "expenses": expenses,
        "flows": flows,
        "inventory": inventory,
        "worksheets": workbook.sheetnames,
    }


def iter_doc_blocks(document: Document):
    for child in document.element.body.iterchildren():
        if isinstance(child, CT_P):
            yield "paragraph", Paragraph(child, document)
        elif isinstance(child, CT_Tbl):
            yield "table", Table(child, document)


def table_to_rows(table: Table) -> list[list[str]]:
    rows: list[list[str]] = []
    for row in table.rows:
        rows.append([cell.text.strip() for cell in row.cells])
    return rows


def is_section_heading(text: str) -> bool:
    if text in {"Premessa", "Allegati"}:
        return True
    return bool(re.match(r"^\d+(?:\.\d+)?\.\s+\S", text))


def read_statute() -> dict[str, Any]:
    document = Document(STATUTE_PATH)
    front_blocks: list[dict[str, Any]] = []
    sections: list[dict[str, Any]] = []
    current: dict[str, Any] | None = None
    in_index = False

    for kind, item in iter_doc_blocks(document):
        if kind == "paragraph":
            text = item.text.strip()
            if not text:
                continue
            text = re.sub(r"\s+", " ", text)

            if text == "Indice sintetico":
                in_index = True
                continue
            if in_index:
                if text == "Allegati":
                    in_index = False
                continue

            if is_section_heading(text):
                current = {"title": text, "blocks": []}
                sections.append(current)
                continue

            block = {"type": "paragraph", "text": text}
            if current is None:
                front_blocks.append(block)
            else:
                current["blocks"].append(block)
        else:
            rows = table_to_rows(item)
            if not rows:
                continue
            block = {"type": "table", "rows": rows}
            if current is None:
                front_blocks.append(block)
            else:
                current["blocks"].append(block)

    front_paragraphs = [block["text"] for block in front_blocks if block["type"] == "paragraph"]
    return {
        "title": front_paragraphs[0] if front_paragraphs else "Statuto del Calottino",
        "subtitle": front_paragraphs[1] if len(front_paragraphs) > 1 else "",
        "notice": front_paragraphs[2] if len(front_paragraphs) > 2 else "",
        "frontBlocks": front_blocks,
        "sections": sections,
        "sectionCount": len(sections),
    }


def build_logo() -> None:
    ASSETS_DIR.mkdir(parents=True, exist_ok=True)
    with Image.open(LOGO_SOURCE_PATH) as image:
        image = image.convert("RGB")
        # The WhatsApp photo is portrait. This crop isolates the circular patch.
        crop_box = (35, 440, 885, 1290)
        patch = image.crop(crop_box)
        patch = patch.resize((720, 720), Image.Resampling.LANCZOS)
        patch.save(ASSETS_DIR / "logo-patch.jpg", quality=92, optimize=True)


def main() -> None:
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    ASSETS_DIR.mkdir(parents=True, exist_ok=True)

    data = {
        "generatedAt": datetime.now(timezone.utc).isoformat(),
        "sources": {
            "excel": {
                "file": EXCEL_PATH.name,
                "modifiedAt": datetime.fromtimestamp(EXCEL_PATH.stat().st_mtime).isoformat(),
            },
            "statute": {
                "file": STATUTE_PATH.name,
                "modifiedAt": datetime.fromtimestamp(STATUTE_PATH.stat().st_mtime).isoformat(),
            },
        },
        "excel": read_workbook(),
        "statute": read_statute(),
    }

    build_logo()
    out_path = DATA_DIR / "site-data.json"
    out_path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Dati sito aggiornati: {out_path}")


if __name__ == "__main__":
    main()
