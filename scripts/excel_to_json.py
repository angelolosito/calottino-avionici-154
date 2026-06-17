#!/usr/bin/env python3
"""Convert the Calottino Excel workbook into the JSON consumed by GitHub Pages."""

from __future__ import annotations

import argparse
import json
import math
import warnings
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Any

try:
    import openpyxl
except ImportError as exc:  # pragma: no cover - used by GitHub Actions and local CLI
    print(f"Modulo mancante: {exc.name}")
    print("Installa le dipendenze con: python3 -m pip install openpyxl")
    raise SystemExit(1)


warnings.filterwarnings("ignore", message="Data Validation extension is not supported")


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_EXCEL = ROOT / "calottino_categoria_gestione.xlsx"
DEFAULT_OUTPUT = ROOT / "docs" / "data.json"
LEGACY_DATA = ROOT / "docs" / "data" / "site-data.json"


def clean_value(value: Any) -> Any:
    if isinstance(value, datetime):
        return value.strftime("%d/%m/%Y")
    if isinstance(value, date):
        return value.strftime("%d/%m/%Y")
    if isinstance(value, float) and value.is_integer():
        return int(value)
    if isinstance(value, str):
        normalized = value.replace("\xa0", " ").strip()
        return normalized if normalized else None
    return value


def present(value: Any) -> bool:
    return value is not None and value != ""


def number(value: Any, fallback: float = 0) -> float:
    value = clean_value(value)
    if value is None:
        return fallback
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        try:
            return float(value.replace(".", "").replace(",", "."))
        except ValueError:
            return fallback
    return fallback


def is_yes(value: Any) -> bool:
    normalized = str(clean_value(value) or "").casefold()
    return normalized in {"si", "sì", "yes", "true", "pagata", "pagato", "approvata", "approvato"}


def table_from_sheet(
    workbook: Any,
    sheet_name: str,
    header_row: int,
    required_headers: list[str] | None = None,
) -> dict[str, Any]:
    sheet = workbook[sheet_name]
    headers = [clean_value(sheet.cell(header_row, col).value) for col in range(1, sheet.max_column + 1)]
    headers = [str(header) for header in headers if present(header)]
    rows: list[dict[str, Any]] = []

    for row_idx in range(header_row + 1, sheet.max_row + 1):
        row: dict[str, Any] = {}
        for col_idx, header in enumerate(headers, start=1):
            row[header] = clean_value(sheet.cell(row_idx, col_idx).value)

        keys = required_headers or headers
        if any(present(row.get(key)) for key in keys):
            rows.append(row)

    return {
        "title": clean_value(sheet.cell(1, 1).value) or sheet_name,
        "sheet": sheet_name,
        "headers": headers,
        "rows": rows,
    }


def table_until_blank(
    sheet: Any,
    title_cell: str,
    header_row: int,
    first_col: int,
    last_col: int,
    data_start: int,
) -> dict[str, Any]:
    headers = [
        str(clean_value(sheet.cell(header_row, col).value))
        for col in range(first_col, last_col + 1)
        if present(clean_value(sheet.cell(header_row, col).value))
    ]
    rows: list[dict[str, Any]] = []

    for row_idx in range(data_start, sheet.max_row + 1):
        row: dict[str, Any] = {}
        for offset, header in enumerate(headers):
            row[header] = clean_value(sheet.cell(row_idx, first_col + offset).value)
        if not any(present(value) for value in row.values()):
            break
        rows.append(row)

    return {
        "title": clean_value(sheet[title_cell].value),
        "headers": headers,
        "rows": rows,
    }


def lookup(rows: list[dict[str, Any]], label: str, value_key: str = "Importo") -> Any:
    for row in rows:
        if row.get("Voce") == label:
            return row.get(value_key)
    return None


def lookup_any(rows: list[dict[str, Any]], labels: list[str], value_key: str = "Importo") -> Any:
    for label in labels:
        value = lookup(rows, label, value_key)
        if value is not None:
            return value
    return None


def parameter_value(rows: list[dict[str, Any]], label: str) -> float:
    for row in rows:
        if row.get("Voce") == label:
            return number(row.get("Valore"))
    return 0


def is_advance_reimbursement(row: dict[str, Any]) -> bool:
    text = " ".join(
        str(clean_value(row.get(key)) or "")
        for key in ("Categoria", "Descrizione", "Note")
    ).casefold()
    return "rimborso" in text and "anticipo" in text


def build_flows(summary: dict[str, Any]) -> dict[str, Any]:
    rows = [
        {
            "Voce": "Quote incassate",
            "Importo": summary["duesCollected"],
            "Segno": "Entrata",
            "Note": "Quote iniziali effettive pagate dai componenti",
        },
    ]

    if summary["advancesReceived"]:
        rows.append(
            {
                "Voce": "Anticipi da restituire",
                "Importo": summary["advancesOutstanding"],
                "Segno": "Debito",
                "Note": "Anticipi temporanei registrati nelle quote, esclusi dagli incassi quota",
            }
        )

    rows.extend(
        [
            {
                "Voce": "Ricavi patch incassati totali",
                "Importo": summary["patchRevenue"],
                "Segno": "Entrata",
                "Note": "Vendite incassate totali: esterni + interni",
            },
            {
                "Voce": "Costo acquisti patch",
                "Importo": summary["patchPurchaseCost"],
                "Segno": "Uscita",
                "Note": "Acquisti patch registrati come pagati/parziali",
            },
            {
                "Voce": "Spese categoria",
                "Importo": summary["categoryExpenses"],
                "Segno": "Uscita",
                "Note": "Spese approvate, inclusi eventuali rimborsi anticipo già pagati",
            },
            {
                "Voce": "Cassa disponibile stimata",
                "Importo": summary["cashAvailable"],
                "Segno": "Saldo",
                "Note": "Entrate, ricavi e anticipi temporanei - uscite",
            },
        ]
    )

    if summary["advancesOutstanding"]:
        rows.append(
            {
                "Voce": "Saldo netto dopo rimborsi",
                "Importo": summary["netAfterAdvances"],
                "Segno": "Saldo",
                "Note": "Cassa stimata al netto degli anticipi ancora da restituire",
            }
        )

    rows.extend(
        [
            {
                "Voce": "Ricavi patch esterni incassati",
                "Importo": summary["patchRevenueExternal"],
                "Segno": "Entrata",
                "Note": "Vendite incassate a prezzo esterni",
            },
            {
                "Voce": "Ricavi patch interni incassati",
                "Importo": summary["patchRevenueInternal"],
                "Segno": "Entrata",
                "Note": "Vendite incassate a prezzo interni",
            },
        ]
    )

    return {
        "title": "Flussi principali",
        "headers": ["Voce", "Importo", "Segno", "Note"],
        "rows": rows,
    }


def build_inventory(summary: dict[str, Any]) -> dict[str, Any]:
    return {
        "title": "Magazzino patch",
        "headers": ["Voce", "Valore", "Unità", "Note"],
        "rows": [
            {
                "Voce": "Patch acquistate",
                "Valore": summary["patchPurchased"],
                "Unità": "pezzi",
                "Note": "Totale inserito in Acquisti_Patch",
            },
            {
                "Voce": "Patch vendute",
                "Valore": summary["patchSold"],
                "Unità": "pezzi",
                "Note": "Totale inserito in Vendite_Patch",
            },
            {
                "Voce": "Patch disponibili",
                "Valore": summary["patchAvailable"],
                "Unità": "pezzi",
                "Note": "Acquistate - vendute",
            },
            {
                "Voce": "Break-even acquisti a prezzo esterni",
                "Valore": summary["breakEvenPatchesExternal"],
                "Unità": "pezzi",
                "Note": "Patch da vendere a prezzo esterni per coprire il costo acquisti",
            },
            {
                "Voce": "Margine unitario esterni",
                "Valore": summary["patchUnitMarginExternal"],
                "Unità": "€",
                "Note": "Prezzo esterni - costo unitario",
            },
            {
                "Voce": "Break-even acquisti a prezzo interni",
                "Valore": summary["breakEvenPatchesInternal"],
                "Unità": "pezzi",
                "Note": "Patch da vendere a prezzo interni per coprire il costo acquisti",
            },
            {
                "Voce": "Margine unitario interni",
                "Valore": summary["patchUnitMarginInternal"],
                "Unità": "€",
                "Note": "Prezzo interni - costo unitario",
            },
        ],
    }


def load_statute(output_path: Path) -> dict[str, Any]:
    for source in (output_path, LEGACY_DATA):
        if source.exists():
            try:
                data = json.loads(source.read_text(encoding="utf-8"))
            except json.JSONDecodeError:
                continue
            statute = data.get("statute")
            if isinstance(statute, dict):
                return statute

    return {
        "title": "Statuto del calottino",
        "subtitle": "",
        "notice": "",
        "frontBlocks": [],
        "sections": [],
        "sectionCount": 0,
    }


def read_workbook(excel_path: Path) -> dict[str, Any]:
    workbook = openpyxl.load_workbook(excel_path, data_only=True)

    people = table_from_sheet(workbook, "Persone_Quote", header_row=3, required_headers=["ID", "Nome / Cognome"])
    purchases = table_from_sheet(
        workbook,
        "Acquisti_Patch",
        header_row=3,
        required_headers=["Data", "Fornitore", "Lotto / Descrizione", "Quantità", "Costo totale", "Pagato da"],
    )
    sales = table_from_sheet(
        workbook,
        "Vendite_Patch",
        header_row=3,
        required_headers=["Data", "Acquirente", "Quantità", "Ricavo"],
    )
    expenses = table_from_sheet(
        workbook,
        "Spese_Categoria",
        header_row=3,
        required_headers=["Data", "Categoria", "Descrizione", "Importo", "Pagato da"],
    )
    parameters = table_from_sheet(workbook, "Parametri", header_row=3, required_headers=["Voce"])

    dashboard_sheet = workbook["Dashboard"]

    member_rows = people["rows"]
    purchase_rows = purchases["rows"]
    sale_rows = sales["rows"]
    expense_rows = expenses["rows"]
    parameter_rows = parameters["rows"]

    paid_members = sum(1 for row in member_rows if row.get("Stato") == "Pagata")
    partial_members = sum(1 for row in member_rows if row.get("Stato") == "Parziale")
    due_total = sum(number(row.get("Quota dovuta")) for row in member_rows)
    paid_total = sum(number(row.get("Quota pagata")) for row in member_rows)
    remaining_total = max(due_total - paid_total, 0)
    advances_received = sum(number(row.get("Anticipo da rimborsare")) for row in member_rows)

    external_price = parameter_value(parameter_rows, "Prezzo rivendita patch Esterni")
    internal_price = parameter_value(parameter_rows, "Prezzo rivendita patch Interni")
    external_margin = parameter_value(parameter_rows, "Margine unitario patch Esterni")
    internal_margin = parameter_value(parameter_rows, "Margine unitario patch Interni")

    patch_purchased = sum(number(row.get("Quantità")) for row in purchase_rows)
    patch_sold = sum(number(row.get("Quantità")) for row in sale_rows)
    patch_available = patch_purchased - patch_sold
    patch_purchase_cost = sum(
        number(row.get("Costo totale"))
        for row in purchase_rows
        if is_yes(row.get("Stato pagamento")) or str(row.get("Stato pagamento") or "").casefold() == "parziale"
    )
    approved_expenses = sum(number(row.get("Importo")) for row in expense_rows if is_yes(row.get("Approvata?")))
    advances_reimbursed = sum(
        number(row.get("Importo"))
        for row in expense_rows
        if is_yes(row.get("Approvata?")) and is_advance_reimbursement(row)
    )
    advances_outstanding = max(advances_received - advances_reimbursed, 0)
    collected_sales = [row for row in sale_rows if is_yes(row.get("Incassato?"))]
    patch_revenue = sum(number(row.get("Ricavo")) for row in collected_sales)
    patch_gross_profit = sum(number(row.get("Utile lordo")) for row in collected_sales)
    patch_revenue_external = sum(
        number(row.get("Ricavo")) for row in collected_sales if number(row.get("Prezzo unitario")) == external_price
    )
    patch_revenue_internal = sum(
        number(row.get("Ricavo")) for row in collected_sales if number(row.get("Prezzo unitario")) == internal_price
    )
    cash_available = paid_total + patch_revenue + advances_received - patch_purchase_cost - approved_expenses
    net_after_advances = cash_available - advances_outstanding

    break_even_external = math.ceil(patch_purchase_cost / external_price) if external_price else 0
    break_even_internal = math.ceil(patch_purchase_cost / internal_price) if internal_price else 0

    summary = {
        "cashAvailable": cash_available,
        "netAfterAdvances": net_after_advances,
        "duesCollected": paid_total,
        "advancesReceived": advances_received,
        "advancesReimbursed": advances_reimbursed,
        "advancesOutstanding": advances_outstanding,
        "patchRevenue": patch_revenue,
        "patchRevenueExternal": patch_revenue_external,
        "patchRevenueInternal": patch_revenue_internal,
        "patchPurchaseCost": patch_purchase_cost,
        "categoryExpenses": approved_expenses,
        "patchGrossProfit": patch_gross_profit,
        "patchPurchased": patch_purchased,
        "patchSold": patch_sold,
        "patchAvailable": patch_available,
        "breakEvenPatches": break_even_external,
        "breakEvenPatchesExternal": break_even_external,
        "breakEvenPatchesInternal": break_even_internal,
        "patchUnitMargin": external_margin,
        "patchUnitMarginExternal": external_margin,
        "patchUnitMarginInternal": internal_margin,
        "membersTotal": len(member_rows),
        "membersPaid": paid_members,
        "membersPartial": partial_members,
        "membersOpen": max(len(member_rows) - paid_members - partial_members, 0),
        "duesExpected": due_total,
        "duesRemaining": remaining_total,
        "duesCompletionRate": round((paid_total / due_total) * 100, 1) if due_total else 0,
        "purchasesCount": len(purchase_rows),
        "salesCount": len(sale_rows),
        "expensesCount": len(expense_rows),
    }

    return {
        "workbookTitle": clean_value(dashboard_sheet["A1"].value),
        "workbookSubtitle": clean_value(dashboard_sheet["A2"].value),
        "summary": summary,
        "parameters": parameters,
        "people": people,
        "purchases": purchases,
        "sales": sales,
        "expenses": expenses,
        "flows": build_flows(summary),
        "inventory": build_inventory(summary),
        "worksheets": workbook.sheetnames,
    }


def build_data(excel_path: Path, output_path: Path) -> dict[str, Any]:
    return {
        "generatedAt": datetime.now(timezone.utc).isoformat(),
        "sources": {
            "excel": {
                "file": excel_path.name,
                "modifiedAt": datetime.fromtimestamp(excel_path.stat().st_mtime).isoformat(),
            }
        },
        "excel": read_workbook(excel_path),
        "statute": load_statute(output_path),
    }


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Converti l'Excel del calottino in docs/data.json.")
    parser.add_argument("--input", type=Path, default=DEFAULT_EXCEL, help="File Excel sorgente.")
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT, help="File JSON generato per il sito.")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    excel_path = args.input.resolve()
    output_path = args.output.resolve()

    if not excel_path.exists():
        raise SystemExit(f"File Excel non trovato: {excel_path}")

    output_path.parent.mkdir(parents=True, exist_ok=True)
    data = build_data(excel_path, output_path)
    output_path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Creato {output_path} da {excel_path}")


if __name__ == "__main__":
    main()
